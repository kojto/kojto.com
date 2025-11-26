#$ kojto_profiles/utils/export_profile_batch_to_excel.py

import io
import base64
from openpyxl import Workbook
import math
from collections import defaultdict
from openpyxl.styles import Font

def export_profile_batch_to_excel(batch):
    """
    Export a kojto.profile.batches record to an Excel file with multiple sheets.
    Args:
        batch: A kojto.profile.batches record (single recordset).
    Returns:
        Tuple of (base64-encoded file content, filename).
    """
    # Create Excel workbook
    output = io.BytesIO()
    workbook = Workbook()

    # Sheet 1: General Information
    sheet1 = workbook.active
    sheet1.title = "General Information"
    sheet1.append(['Field', 'Value'])
    sheet1.append(['Batch Name', batch.name or ''])
    sheet1.append(['Subcode', batch.subcode_id.name or '' if batch.subcode_id else ''])
    sheet1.append(['Subcode Description', batch.subcode_description or ''])
    sheet1.append(['Description', batch.description or ''])
    sheet1.append(['Issue Date', batch.date_issue.strftime('%Y-%m-%d') if batch.date_issue else ''])
    sheet1.append(['Created By', batch.issued_by.name or '' if batch.issued_by else ''])
    sheet1.append(['Total Process Time (hrs)', batch.total_batch_process_time or 0.0])

    # Sheet 2: Batch Content
    sheet2 = workbook.create_sheet(title="Batch Content")
    headers2 = [
        '№', 'Profile', 'Length (m)', 'Length Extension (m)', 'Pcs',
        'Material', 'Profile Weight (kg/m)', 'Coating Perimeter (mm)',
        'Total Net Length (m)', 'Total Gross Length (m)',
        'Total Net Weight (kg)', 'Total Gross Weight (kg)', 'Total Coating Area (m²)',
        'Ext. Corners', 'Total Process Time (min)'
    ]
    sheet2.append(headers2)
    for content in batch.batch_content_ids:
        row2 = [
            content.position or '',
            content.profile_id.name or '',
            content.length or 0.0,
            content.length_extension or 0.0,
            content.quantity or 0,
            content.material_id.name or '' if content.material_id else '',
            content.profile_weight or 0.0,
            content.coating_perimeter or 0.0,
            content.total_profile_length_net or 0.0,
            content.total_profile_length_gross or 0.0,
            content.total_profile_weight_net or 0.0,
            content.total_profile_weight_gross or 0.0,
            content.total_profile_coating_area or 0.0,
            content.number_ext_corners or 0,
            content.total_profile_process_time or 0.0
        ]
        sheet2.append(row2)

    # Sheet 3: Strips
    sheet3 = workbook.create_sheet(title="Strips")
    headers3 = [
        'Position', 'Profile', 'Length (m)', 'Length Ext. (m)', 'Pcs',
        'Strip Name', 'Thickness (mm)', 'Width (mm)', 'Material', 'Unit Weight (kg/m)', 'Total Weight (kg)'
    ]
    sheet3.append(headers3)
    for content in batch.batch_content_ids:
        profile = content.profile_id
        if hasattr(profile, 'strip_ids') and profile.strip_ids:
            for strip in profile.strip_ids:
                total_length = content.length + content.length_extension
                # Calculate unit weight (kg/m) = strip_cross_sectional_area * material_density * 100
                material_density = content.material_id.density if content.material_id and hasattr(content.material_id, 'density') else 1000  # kg/m³
                unit_weight = (strip.strip_cross_sectional_area or 0.0) * material_density * 100  # Convert cm² to m²
                # Calculate total weight (kg) = unit_weight * total_length * quantity
                total_weight = unit_weight * total_length * (content.quantity or 0)

                for _ in range(int(content.quantity)):
                    row3 = [
                        content.position or '',
                        profile.name or '',
                        content.length or 0.0,
                        content.length_extension or 0.0,
                        content.quantity or 0,
                        strip.name or '',
                        strip.thickness or 0.0,
                        strip.projected_length or 0.0,
                        content.material_id.name or '' if content.material_id else '',
                        round(unit_weight, 3) if unit_weight else 0.0,
                        round(total_weight, 3) if total_weight else 0.0
                    ]
                    sheet3.append(row3)

    # Sheet 4: Consolidated Strips by Material + Thickness
    strip_data = defaultdict(lambda: defaultdict(int))
    for content in batch.batch_content_ids:
        profile = content.profile_id
        if hasattr(profile, 'strip_ids') and profile.strip_ids:
            for strip in profile.strip_ids:
                material = content.material_id.name if content.material_id else ''
                thickness = strip.thickness if strip.thickness else 0.0
                total_length = content.length + content.length_extension
                projected_length = strip.projected_length or 0.0
                pcs = content.quantity or 0
                key = (material, thickness, projected_length, total_length)
                strip_data[(material, thickness)][(projected_length, total_length)] += pcs

    for (material, thickness), length_data in strip_data.items():
        sheet_name = f"{thickness}mm_{material or 'NoMaterial'}"[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        headers = ['Strip Width (mm)', 'Strip Length (mm)', 'Pcs', 'Unit Weight (kg)', 'Total Weight (kg)']
        sheet.append(headers)

        # Get material density for this material
        material_density = 1000  # Default density
        if material:
            # Try to find the material and get its density
            material_record = batch.env['kojto.base.material.grades'].search([('name', '=', material)], limit=1)
            if material_record and hasattr(material_record, 'density'):
                material_density = material_record.density

        for (projected_length, total_length), pcs in length_data.items():
            # Calculate unit weight: width × length × thickness × density / 1,000,000,000
            # Convert mm to m: (width_mm × length_mm × thickness_mm × density_kg_m3) / 1,000,000,000
            unit_weight = (projected_length * total_length * thickness * material_density) / 1000000000
            total_weight = unit_weight * pcs

            row = [projected_length, total_length, pcs, round(unit_weight, 6), round(total_weight, 6)]
            sheet.append(row)

    # Sheet 5: Processes
    sheet5 = workbook.create_sheet(title="Processes")
    headers5 = [
        'Profile', 'Process Name', 'Process Type', 'Time per Meter (min/m)', 'Description'
    ]
    sheet5.append(headers5)
    seen_profiles = set()
    for content in batch.batch_content_ids:
        profile = content.profile_id
        if profile.id not in seen_profiles and hasattr(profile, 'process_ids') and profile.process_ids:
            seen_profiles.add(profile.id)
            for process in profile.process_ids:
                row5 = [
                    profile.name or '',
                    process.description or '',
                    process.process_type or '',
                    process.time_per_meter or 0.0,
                    process.description or ''
                ]
                sheet5.append(row5)

    # Sheet 6: Process Type Consolidation
    sheet6 = workbook.create_sheet(title="Process Type Consolidation")
    headers6 = ['Process Type', 'Total Time (hrs)']  # Removed 'Number of Profiles'
    sheet6.append(headers6)

    # Consolidate process times by process_type
    process_type_data = defaultdict(lambda: {'total_time': 0.0})  # Removed 'profile_count'
    for content in batch.batch_content_ids:
        profile = content.profile_id
        if hasattr(profile, 'process_ids') and profile.process_ids:
            gross_length_per_piece = (content.length + content.length_extension) / 1000  # Convert mm to m
            total_length = gross_length_per_piece * (content.quantity or 0)  # Total length in meters
            for process in profile.process_ids:
                process_type = process.process_type or 'Unknown'
                time_per_meter = process.time_per_meter or 0.0
                process_time = total_length * time_per_meter  # Time in minutes
                process_type_data[process_type]['total_time'] += process_time

    # Write consolidated data to sheet, converting minutes to hours and rounding to 2 decimals
    for process_type, data in process_type_data.items():
        total_time_hours = round(data['total_time'] / 60, 2)  # Convert to hours and round to 2 decimals
        row6 = [
            process_type,
            total_time_hours
        ]
        sheet6.append(row6)

    # Sheet 7: Shapes
    sheet7 = workbook.create_sheet(title="Shapes")
    headers7 = [
        'Position', 'Profile', 'Shape Name', 'Length (m)', 'Length Ext. (m)', 'Pcs', 'Material', 'Insert Count'
    ]
    sheet7.append(headers7)

    # Use SQL to fetch shapes and count insertions per profile
    shape_insert_counts = defaultdict(lambda: defaultdict(int))
    profile_shapes = defaultdict(list)
    for content in batch.batch_content_ids:
        profile_id = content.profile_id.id if content.profile_id else None
        if profile_id:
            # Fetch shapes linked to this profile via shape inserts
            batch.env.cr.execute("""
                SELECT si.shape_id, COUNT(si.id) as insert_count
                FROM kojto_profile_shape_inserts si
                WHERE si.profile_id = %s AND si.shape_id IS NOT NULL
                GROUP BY si.shape_id
            """, (profile_id,))
            results = batch.env.cr.fetchall()
            for shape_id, insert_count in results:
                shape_insert_counts[profile_id][shape_id] = insert_count
                if shape_id not in profile_shapes[profile_id]:
                    profile_shapes[profile_id].append(shape_id)

            # Now get shape details
            for shape_id in profile_shapes[profile_id]:
                shape = batch.env['kojto.profile.shapes'].browse(shape_id)
                if shape.exists():
                    total_pcs = (content.quantity or 0) * shape_insert_counts[profile_id][shape_id]
                    row7 = [
                        content.position or '',
                        content.profile_id.name or '',
                        shape.name or '',
                        content.length or 0.0,
                        content.length_extension or 0.0,
                        total_pcs,
                        content.material_id.name or '' if content.material_id else '',
                        shape_insert_counts[profile_id][shape_id]
                    ]
                    sheet7.append(row7)

    # Sheet 8: Individual Shape Sheets
    shape_data = defaultdict(list)
    shape_names = {}
    for content in batch.batch_content_ids:
        profile_id = content.profile_id.id if content.profile_id else None
        if profile_id and profile_id in profile_shapes:
            for shape_id in profile_shapes[profile_id]:
                shape = batch.env['kojto.profile.shapes'].browse(shape_id)
                if shape.exists():
                    shape_name = shape.name or 'UnknownShape'
                    shape_names[shape_id] = shape_name
                    total_pcs = (content.quantity or 0) * shape_insert_counts[profile_id][shape_id]
                    shape_data[shape_id].append({
                        'position': content.position or '',
                        'profile': content.profile_id.name or '',
                        'length': content.length or 0.0,
                        'length_ext': content.length_extension or 0.0,
                        'pcs': total_pcs,
                        'material': content.material_id.name or '' if content.material_id else '',
                        'insert_count': shape_insert_counts[profile_id][shape_id]
                    })

    # Individual sheets per shape
    for shape_id, shape_data in shape_data.items():
        shape_name = shape_names.get(shape_id, 'UnknownShape')
        # Ensure sheet name is within Excel's 31 character limit
        sheet_name = f"SHP_{shape_name}"[:31]
        if sheet_name in workbook.sheetnames:
            shape_sheet = workbook[sheet_name]
        else:
            shape_sheet = workbook.create_sheet(sheet_name)
            # Set headers
            headers = ['Position', 'Length', 'Material', 'Pcs']
            for col, header in enumerate(headers, 1):
                shape_sheet.cell(row=1, column=col).value = header
                shape_sheet.cell(row=1, column=col).font = Font(bold=True)

        # Consolidate data by length and material
        consolidated_data = defaultdict(lambda: {'length': 0, 'pcs': 0, 'positions': set()})
        for data in shape_data:
            key = (data['length'], data['material'])
            consolidated_data[key]['length'] = data['length']
            consolidated_data[key]['pcs'] += data['pcs']
            consolidated_data[key]['positions'].add(data['position'])

        # Write consolidated data to sheet with new position numbers
        row_idx = 2
        position_num = 1
        for key, data in consolidated_data.items():
            shape_sheet.cell(row=row_idx, column=1).value = position_num
            shape_sheet.cell(row=row_idx, column=2).value = data['length']
            shape_sheet.cell(row=row_idx, column=3).value = key[1]  # Material
            shape_sheet.cell(row=row_idx, column=4).value = data['pcs']
            row_idx += 1
            position_num += 1

        # Adjust column widths
        for col in shape_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            shape_sheet.column_dimensions[column].width = adjusted_width

    # Save workbook
    workbook.save(output)
    output.seek(0)
    file_content = output.read()
    output.close()

    # Return base64-encoded content and filename
    filename = f"Profile_Batch_{batch.name or 'Export'}.xlsx"
    return base64.b64encode(file_content), filename
