#$ kojto_optimizer/utils/export_1d_cutting_plan_to_excel.py

import json
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_1d_cutting_plan_to_excel(record):
    """Export the cutting_plan_json to an Excel file with Summary, Cutting Plans Rel, and Cutting Plans Abs sheets.
    Cutting Plans Rel includes Cut 0; Cutting Plans Abs uses static values with Waste as last column.
    No width_of_cut added after Cut 0 if initial_cut is 0."""
    if not record.cutting_plan_json:
        raise ValidationError("No cutting plan available to export.")

    # Parse the JSON
    try:
        cutting_plan_data = json.loads(record.cutting_plan_json)
    except json.JSONDecodeError:
        raise ValidationError("Invalid cutting plan JSON data.")

    # Create a new Excel workbook
    wb = Workbook()

    # First sheet: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Summary headers and package data
    package_data = [
        ("ID", cutting_plan_data.get("id", "")),
        ("Name", cutting_plan_data.get("name", "")),
        ("Subcode ID", cutting_plan_data.get("subcode_id", "")),
        ("Description", cutting_plan_data.get("description", "")),
        ("Date Issue", cutting_plan_data.get("date_issue", "")),
    ]
    row = 1
    for label, value in package_data:
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    # Add a blank row
    row += 1

    # Summary metrics, including width_of_cut
    summary = cutting_plan_data.get("summary", {})
    summary_metrics = [
        ("Available Stock Length (m)", summary.get("total_stock_length", 0.0) / 1000),
        ("Used Stock Length (m)", summary.get("total_used_stock_length", 0.0) / 1000),
        ("Total Bar Length (m)", summary.get("total_bar_length", 0.0) / 1000),
        ("Total Waste (%)", summary.get("total_waste_percentage", 0.0)),
        ("Initial Cut (mm)", record.initial_cut),
        ("Width of Cut (mm)", record.width_of_cut),
        ("Final Cut (mm)", record.final_cut),
    ]
    for label, value in summary_metrics:
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    # Add Stock Used section
    row += 1
    ws_summary.cell(row=row, column=1, value="Stock Used")
    row += 1
    stock_headers = ["Stock Position", "Stock Description", "Stock Length (mm)", "Pieces"]
    for col, header in enumerate(stock_headers, start=1):
        ws_summary.cell(row=row, column=col, value=header)
    row += 1

    for stock in cutting_plan_data.get("stock_used", []):
        ws_summary.append([
            stock["stock_position"],
            stock["stock_description"],
            stock["stock_length"],
            stock["pcs"],
        ])
        row += 1

    # Add message
    row += 1
    ws_summary.cell(row=row, column=1, value="Message")
    ws_summary.cell(row=row, column=2, value=cutting_plan_data.get("message", ""))

    # Adjust column widths for Summary sheet
    for col in range(1, 5):
        column_letter = get_column_letter(col)
        ws_summary.column_dimensions[column_letter].width = 20

    # Second sheet: Cutting Plans Rel (Relative values, including Cut 0)
    ws_cutting_rel = wb.create_sheet(title="Cutting Plans Rel")

    # Determine the maximum number of cuts for dynamic column sizing
    max_cuts = max(len(plan["cut_pattern"]) for plan in cutting_plan_data.get("cutting_plans", [])) if cutting_plan_data.get("cutting_plans") else 0

    # Headers for Cutting Plans Rel (add Cut 0)
    headers_rel = [
        "Stock ID", "Stock Position", "Stock Description", "Stock Length (mm)",
        "Pieces", "Waste (%)", "Cut 0 (mm)"
    ]
    headers_rel.extend([f"Cut {i+1} (mm)" for i in range(max_cuts)])

    ws_cutting_rel.append(headers_rel)

    # Populate Cutting Plans Rel with split cuts, including Cut 0
    for plan in cutting_plan_data.get("cutting_plans", []):
        cut_lengths = [cut["length"] for cut in plan["cut_pattern"]]
        row_data = [
            plan["stock_id"],
            plan["stock_position"],
            plan["stock_description"],
            plan["stock_length"],
            plan["pieces"],
            plan["waste_percentage"],
            record.initial_cut,  # Cut 0
        ]
        row_data.extend(cut_lengths)
        row_data.extend([""] * (max_cuts - len(cut_lengths)))
        ws_cutting_rel.append(row_data)

    # Adjust column widths for Cutting Plans Rel sheet
    for col in range(1, len(headers_rel) + 1):
        column_letter = get_column_letter(col)
        ws_cutting_rel.column_dimensions[column_letter].width = 15

    # Third sheet: Cutting Plans Abs (Absolute values, static, with Waste)
    ws_cutting_abs = wb.create_sheet(title="Cutting Plans Abs")

    # Headers for Cutting Plans Abs (rename last column to Waste)
    headers_abs = [
        "Stock ID", "Stock Position", "Stock Description", "Stock Length (mm)",
        "Pieces", "Cut 0 (mm)"
    ]
    headers_abs.extend([f"Cut {i+1} (mm)" for i in range(max_cuts)])
    headers_abs.append("Waste (mm)")

    ws_cutting_abs.append(headers_abs)

    # Populate Cutting Plans Abs with static values
    for i, plan in enumerate(cutting_plan_data.get("cutting_plans", []), start=2):  # Start at row 2 (after header)
        cut_lengths = [cut["length"] for cut in plan["cut_pattern"]]

        # Calculate absolute cumulative values
        absolute_lengths = []
        cumulative = record.initial_cut  # Cut 0, no width_of_cut if initial_cut is 0
        if record.initial_cut > 0:
            cumulative += record.width_of_cut  # Add width_of_cut only if initial_cut > 0
        absolute_lengths.append(cumulative)

        for length in cut_lengths:
            cumulative += length + record.width_of_cut
            absolute_lengths.append(cumulative)

        # Calculate waste as stock_length - last cut
        last_cut = absolute_lengths[-1] if absolute_lengths else cumulative
        waste = plan["stock_length"] - last_cut

        # Base row data
        ws_cutting_abs.cell(row=i, column=1, value=plan["stock_id"])
        ws_cutting_abs.cell(row=i, column=2, value=plan["stock_position"])
        ws_cutting_abs.cell(row=i, column=3, value=plan["stock_description"])
        ws_cutting_abs.cell(row=i, column=4, value=plan["stock_length"])
        ws_cutting_abs.cell(row=i, column=5, value=plan["pieces"])

        # Add static absolute values
        for j, abs_value in enumerate(absolute_lengths):
            ws_cutting_abs.cell(row=i, column=6 + j, value=abs_value)

        # Pad with empty cells if needed
        for j in range(len(cut_lengths), max_cuts):
            ws_cutting_abs.cell(row=i, column=7 + j, value="")

        # Waste column
        ws_cutting_abs.cell(row=i, column=7 + max_cuts, value=waste)

    # Adjust column widths for Cutting Plans Abs sheet
    for col in range(1, len(headers_abs) + 1):
        column_letter = get_column_letter(col)
        ws_cutting_abs.column_dimensions[column_letter].width = 15

    # Save to a BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create attachment and return action to download
    filename = f"{record.name}.xlsx"
    attachment = record.env['ir.attachment'].create({
        'name': filename,
        'datas': base64.encodebytes(output.getvalue()),
        'res_model': record._name,
        'res_id': record.id,
        'type': 'binary',
    })

    return {
        'type': 'ir.actions.act_url',
        'url': f'/web/content/{attachment.id}?download=true',
        'target': 'self',
    }
